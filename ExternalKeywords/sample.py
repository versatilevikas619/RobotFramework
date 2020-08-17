import openpyxl
import json

wb=openpyxl.load_workbook(r"C:\Users\vikas\Desktop\Book1.xlsx")

def FindMaxRow(SheetName):
    y=wb[SheetName]
    row=y.max_row
    return row

def ReadData(SheetName, row, col):
    sh=wb[SheetName]
    cell=sh.cell(int(row),int(col))
    return cell.value




